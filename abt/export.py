"""
abt/export.py
─────────────────────────────────────────────────────────────────────────────
Converts an ABTProfile and the run_analysis() result dict into an xlsx file.
One row per column. Frozen header row, auto column widths, colour-coded
readiness and health columns so the file is readable immediately in Excel.

Requires: openpyxl  (pip install openpyxl)
"""

import io
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .columnProfile import ABTProfile

# ── Column definitions: (header, field_key, width) ───────────────────────────

COLUMNS = [
    # Table-level
    ("table_name",           "table_name",           20),
    ("version",              "version",              8),
    ("snapshot_date",        "snapshot_date",        14),
    ("row_count",            "row_count",            10),
    # Column identity
    ("column_name",          "column_name",          22),
    ("data_type",            "data_type",            10),
    ("statistical_scale",    "statistical_scale",    16),
    ("ordinal_position",     "ordinal_position",     10),
    # Completeness
    ("completeness_pct",     "completeness_pct",     14),
    ("missing_count",        "missing_count",        12),
    # Cardinality
    ("cardinality",          "cardinality",          12),
    ("uniqueness_pct",       "uniqueness_pct",       13),
    # Numeric stats
    ("mean",                 "mean",                 12),
    ("median",               "median",               12),
    ("std",                  "std",                  12),
    ("min",                  "min",                  12),
    ("max",                  "max",                  12),
    ("skewness",             "skewness",             10),
    ("kurtosis",             "kurtosis",             10),
    # Quantiles
    ("q25",                  "q25",                  10),
    ("q50",                  "q50",                  10),
    ("q75",                  "q75",                  10),
    # Outliers
    ("has_outliers",         "has_outliers",         12),
    ("n_outliers",           "n_outliers",           10),
    # Data quality
    ("blank_value_count",    "blank_value_count",    16),
    ("mismatched_count",     "mismatched_count",     15),
    # Governance
    ("information_privacy",  "information_privacy",  18),
    ("has_unique_field",     "has_unique_field",     14),
    # Analysis results
    ("readiness_status",     "readiness_status",     16),
    ("health_score",         "health_score",         12),
    ("health_label",         "health_label",         12),
    ("health_top_issue",     "health_top_issue",     40),
    ("is_blocker",           "is_blocker",           10),
    ("is_governance_risk",   "is_governance_risk",   16),
    ("governance_risk_types","governance_risk_types",20),
]

# ── Styles ────────────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1E2336")
HDR_FONT = Font(name="Arial", bold=True, color="E2E8F0", size=10)

_STATUS_FILL = {
    "ready":   PatternFill("solid", fgColor="14532D"),
    "caution": PatternFill("solid", fgColor="451A03"),
    "drop":    PatternFill("solid", fgColor="450A0A"),
}
_STATUS_FONT = {
    "ready":   Font(name="Arial", color="22C55E", bold=True, size=10),
    "caution": Font(name="Arial", color="F59E0B", bold=True, size=10),
    "drop":    Font(name="Arial", color="EF4444", bold=True, size=10),
}

_HEALTH_FILL = {
    "good":     PatternFill("solid", fgColor="14532D"),
    "fair":     PatternFill("solid", fgColor="1E3A5F"),
    "poor":     PatternFill("solid", fgColor="451A03"),
    "critical": PatternFill("solid", fgColor="450A0A"),
}
_HEALTH_FONT = {
    "good":     Font(name="Arial", color="22C55E", bold=True, size=10),
    "fair":     Font(name="Arial", color="3B82F6", bold=True, size=10),
    "poor":     Font(name="Arial", color="F59E0B", bold=True, size=10),
    "critical": Font(name="Arial", color="EF4444", bold=True, size=10),
}

_DEFAULT_FONT  = Font(name="Arial", size=10)
_DEFAULT_ALIGN = Alignment(vertical="center")
_THIN_BORDER   = Border(
    bottom=Side(style="thin", color="2A3050"),
    right= Side(style="thin", color="2A3050"),
)
_CENTER = Alignment(horizontal="center", vertical="center")


def export_analysis_xlsx(abt: ABTProfile, results: Dict) -> bytes:
    """
    Build an xlsx workbook from an ABTProfile and run_analysis() results.
    Returns raw bytes ready for Flask Response().
    """

    # Pre-index result sections for O(1) lookup
    readiness_map = {r["column"]: r["status"] for r in results.get("s5", [])}
    health_map    = results.get("s8", {})
    blocker_set   = {item["column"] for item in results.get("s2", [])}
    governance_map: Dict[str, list] = {
        item["column"]: [r["risk_type"] for r in item.get("risks", [])]
        for item in results.get("s4", [])
    }

    wb = Workbook()
    ws = wb.active
    ws.title       = "Analysis"
    ws.freeze_panes = "A2"

    headers = [c[0] for c in COLUMNS]

    # Header row
    ws.row_dimensions[1].height = 22
    for ci, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = _CENTER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Data rows
    for ri, col in enumerate(abt.columns, 2):
        hs         = health_map.get(col.name, {})
        gov_risks  = governance_map.get(col.name, [])
        readiness  = readiness_map.get(col.name, "")
        health_lbl = hs.get("label", "")

        row_data = {
            "table_name":           abt.abt_name,
            "version":              abt.version,
            "snapshot_date":        abt.snapshot_date,
            "row_count":            abt.row_count,
            "column_name":          col.name,
            "data_type":            col.data_type,
            "statistical_scale":    col.statistical_scale,
            "ordinal_position":     col.ordinal_position,
            "completeness_pct":     col.completeness_percent,
            "missing_count":        col.missing_count,
            "cardinality":          col.cardinality_count,
            "uniqueness_pct":       col.uniqueness_percent,
            "mean":                 col.mean,
            "median":               col.median,
            "std":                  col.std,
            "min":                  col.min_val,
            "max":                  col.max_val,
            "skewness":             col.skewness,
            "kurtosis":             col.kurtosis,
            "q25":                  col.q25,
            "q50":                  col.q50,
            "q75":                  col.q75,
            "has_outliers":         col.has_outliers,
            "n_outliers":           col.n_outliers,
            "blank_value_count":    col.blank_value_count,
            "mismatched_count":     col.mismatched_count,
            "information_privacy":  col.information_privacy or "",
            "has_unique_field":     col.has_unique_field,
            "readiness_status":     readiness,
            "health_score":         hs.get("score", ""),
            "health_label":         health_lbl,
            "health_top_issue":     hs.get("top_issue", ""),
            "is_blocker":           col.name in blocker_set,
            "is_governance_risk":   bool(gov_risks),
            "governance_risk_types": ",".join(gov_risks),
        }

        ws.row_dimensions[ri].height = 18

        for ci, (header, _, _) in enumerate(COLUMNS, 1):
            value = row_data[header]
            cell  = ws.cell(row=ri, column=ci, value=value)
            cell.font      = _DEFAULT_FONT
            cell.alignment = _DEFAULT_ALIGN
            cell.border    = _THIN_BORDER

            if header == "readiness_status" and readiness in _STATUS_FILL:
                cell.fill      = _STATUS_FILL[readiness]
                cell.font      = _STATUS_FONT[readiness]
                cell.alignment = _CENTER

            elif header == "health_label" and health_lbl in _HEALTH_FILL:
                cell.fill      = _HEALTH_FILL[health_lbl]
                cell.font      = _HEALTH_FONT[health_lbl]
                cell.alignment = _CENTER

            elif header == "health_score" and isinstance(value, (int, float)):
                cell.number_format = "0.0"

            elif header in ("completeness_pct", "uniqueness_pct") and isinstance(value, (int, float)):
                cell.number_format = "0.00"

            elif header in ("is_blocker", "is_governance_risk", "has_outliers", "has_unique_field"):
                cell.alignment = _CENTER

    # Column widths
    for ci, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()